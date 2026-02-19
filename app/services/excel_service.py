"""
excel_service.py — 대사 결과를 서식 있는 Excel 파일로 출력.

시트 구성:
  Summary     — 주석별 일치율 요약
  Note_01..N  — 주석별 상세 대사 결과 (다차원 속성 wide format)
  Mismatches  — 불일치 + 미발견 항목 모음
  Mapping_Log — 주석 단위 매핑 신뢰도 로그
"""
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from app.models.reconcile_model import AmountMatch, ReconcileResult
from app.services.mapping_service import NoteMapping

logger = logging.getLogger(__name__)

# FS 타입 → 한글 시트명
_FS_SHEET_NAMES: dict[str, str] = {
    "balance_sheet":    "FS_재무상태표",
    "income_statement": "FS_손익계산서",
    "equity_changes":   "FS_자본변동표",
    "cash_flow":        "FS_현금흐름표",
}

# ─── 색상 ───────────────────────────────────────────────────────
C_MATCH      = "C6EFCE"   # 연초록 — 일치
C_MISMATCH   = "FFC7CE"   # 연빨강 — 불일치
C_LOW_CONF   = "FFEB9C"   # 연노랑 — 신뢰도 낮음 (< 0.8)
C_NOT_FOUND  = "D9D9D9"   # 회색 — 영문 미발견
C_HEADER     = "4472C4"   # 파랑 — 헤더
C_KR_COL     = "DCE6F1"   # 연청색 — 국문 열
C_EN_COL     = "E2EFDA"   # 연녹색 — 영문 열
C_ATTR       = "FFF2CC"   # 연노랑 — 속성 열
C_SUMMARY_OK = "E2EFDA"   # Summary 일치율 높음
C_SUMMARY_NG = "FFC7CE"   # Summary 일치율 낮음


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size, name="맑은 고딕")


def _border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _align(horizontal="left", wrap=False) -> Alignment:
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)


def _write_cell(ws, row: int, col: int, value, fill=None, font=None, align=None, border=True, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    else:
        cell.alignment = _align()
    if border:
        cell.border = _border()
    if number_format:
        cell.number_format = number_format
    return cell


# ─── 공개 API ────────────────────────────────────────────────────

async def generate_excel(
    results: list[ReconcileResult],
    mappings: list[NoteMapping],
    company_name: str,
    output_dir: Path,
    stmt_results: list[ReconcileResult] | None = None,
    stmt_mappings: list[NoteMapping] | None = None,
) -> Path:
    """
    대사 결과를 Excel 파일로 저장.
    stmt_results/stmt_mappings: 재무제표 본문 대사 결과 (없으면 기존 주석만)
    반환: 저장된 파일 경로
    """
    stmt_results  = stmt_results  or []
    stmt_mappings = stmt_mappings or []

    wb = Workbook()
    wb.remove(wb.active)  # 기본 Sheet 제거

    _write_summary(wb, results, mappings, company_name, stmt_results, stmt_mappings)
    _write_mapping_log(wb, mappings, stmt_mappings)
    _write_mismatches(wb, results + stmt_results)

    # FS 시트 (Summary 직후, 주석 시트 전)
    for result in stmt_results:
        sheet_name = _FS_SHEET_NAMES.get(result.note_number_kr, f"FS_{result.note_number_kr}")
        _write_note_sheet(wb, result, sheet_name=sheet_name)

    for result in results:
        _write_note_sheet(wb, result)

    # 파일명: reconciliation_{회사명}_{날짜}.xlsx
    date_str = datetime.now().strftime("%Y%m%d")
    safe_name = company_name.replace(" ", "_").replace("/", "_")[:20]
    filename = f"reconciliation_{safe_name}_{date_str}.xlsx"
    output_path = output_dir / filename

    wb.save(str(output_path))
    logger.info("Excel 저장 완료: %s", output_path)
    return output_path


# ─── Summary 시트 ────────────────────────────────────────────────

def _write_summary(
    wb: Workbook,
    results: list[ReconcileResult],
    mappings: list[NoteMapping],
    company_name: str,
    stmt_results: list[ReconcileResult] | None = None,
    stmt_mappings: list[NoteMapping] | None = None,
):
    stmt_results  = stmt_results  or []
    stmt_mappings = stmt_mappings or []

    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    # 제목
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = f"재무제표 국문/영문 대사 결과 — {company_name}"
    c.font = _font(bold=True, color="FFFFFF", size=13)
    c.fill = _fill(C_HEADER)
    c.alignment = _align("center")

    ws.merge_cells("A2:I2")
    ws["A2"].value = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = _font(size=9, color="666666")
    ws["A2"].alignment = _align("right")

    # 전체 집계 (주석 + FS)
    all_results = stmt_results + results
    total_amounts = sum(r.total_amounts for r in all_results)
    total_matched = sum(r.matched_count for r in all_results)
    total_miss    = sum(
        sum(1 for i in r.items for m in i.amount_matches if m.is_match is False)
        for r in all_results
    )
    total_nf      = sum(
        sum(1 for i in r.items for m in i.amount_matches if not m.found)
        for r in all_results
    )
    overall_rate  = total_matched / total_amounts if total_amounts else 0

    ws.merge_cells("A3:I3")
    ws["A3"].value = (
        f"전체: {total_amounts}건 대사 | 일치 {total_matched}건 ({overall_rate:.1%}) | "
        f"불일치 {total_miss}건 | 미발견 {total_nf}건"
    )
    ws["A3"].font = _font(bold=True, size=11)
    ws["A3"].alignment = _align("center")
    ws["A3"].fill = _fill("F2F2F2")

    # 헤더
    headers = ["번호/구분", "국문 제목", "영문 제목",
               "총 금액수", "일치", "불일치", "미발견", "일치율", "매핑방법"]
    header_row = 5

    def _write_section_header(ws, row: int, label: str):
        ws.merge_cells(f"A{row}:I{row}")
        c = ws[f"A{row}"]
        c.value = label
        c.font = _font(bold=True, color="FFFFFF")
        c.fill = _fill("455A64")
        c.alignment = _align("left")

    def _write_data_rows(ws, result_list, mapping_list, start_row: int) -> int:
        mapping_by_num = {m.kr_note.note_number: m for m in mapping_list}
        row_idx = start_row
        for r in result_list:
            miss  = sum(1 for i in r.items for m in i.amount_matches if m.is_match is False)
            nf    = sum(1 for i in r.items for m in i.amount_matches if not m.found)
            rate  = r.match_rate
            mapping = mapping_by_num.get(r.note_number_kr)
            method  = mapping.method if mapping else "-"
            row_fill = _fill(C_SUMMARY_NG if (miss > 0 or nf > 0) else "FFFFFF")

            vals = [
                r.note_number_kr,
                r.note_title_kr,
                r.note_title_en or "—",
                r.total_amounts,
                r.matched_count,
                miss,
                nf,
                None,
                method,
            ]
            for col, val in enumerate(vals, 1):
                cell = _write_cell(ws, row_idx, col, val, fill=row_fill)
                if col == 1:
                    cell.alignment = _align("center")
                if col == 4:
                    cell.alignment = _align("right")

            miss_cell = ws.cell(row=row_idx, column=6)
            if miss > 0:
                miss_cell.fill = _fill(C_MISMATCH)
                miss_cell.font = _font(bold=True)

            rate_cell = ws.cell(row=row_idx, column=8, value=rate)
            rate_cell.number_format = "0.0%"
            rate_cell.fill = row_fill
            rate_cell.border = _border()
            rate_cell.alignment = _align("center")

            row_idx += 1
        return row_idx

    for col, h in enumerate(headers, 1):
        _write_cell(ws, header_row, col, h,
                    fill=_fill(C_HEADER),
                    font=_font(bold=True, color="FFFFFF"),
                    align=_align("center"))

    next_row = header_row + 1

    # FS 섹션 (있는 경우)
    if stmt_results:
        _write_section_header(ws, next_row, "■ 재무제표 본문")
        next_row += 1
        next_row = _write_data_rows(ws, stmt_results, stmt_mappings, next_row)

    # 주석 섹션
    if results:
        _write_section_header(ws, next_row, "■ 주석 (Notes)")
        next_row += 1
        _write_data_rows(ws, results, mappings, next_row)

    # 열 너비
    col_widths = [12, 28, 35, 8, 8, 8, 8, 8, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A6"


# ─── Mapping_Log 시트 ────────────────────────────────────────────

def _write_mapping_log(
    wb: Workbook,
    mappings: list[NoteMapping],
    stmt_mappings: list[NoteMapping] | None = None,
):
    ws = wb.create_sheet("Mapping_Log")
    ws.sheet_view.showGridLines = False

    headers = ["번호/구분", "국문 제목", "영문 번호/구분", "영문 제목", "매핑방법", "신뢰도"]
    for col, h in enumerate(headers, 1):
        _write_cell(ws, 1, col, h,
                    fill=_fill(C_HEADER),
                    font=_font(bold=True, color="FFFFFF"),
                    align=_align("center"))

    all_mappings = (stmt_mappings or []) + mappings
    for row, m in enumerate(all_mappings, 2):
        en_num   = m.en_note.note_number if m.en_note else "—"
        en_title = m.en_note.note_title  if m.en_note else "영문 미존재"
        conf_fill = _fill(C_MATCH) if m.confidence >= 0.9 else (
                    _fill(C_LOW_CONF) if m.confidence >= 0.5 else _fill(C_MISMATCH))

        vals = [m.kr_note.note_number, m.kr_note.note_title, en_num, en_title, m.method, None]
        for col, val in enumerate(vals, 1):
            _write_cell(ws, row, col, val)

        conf_cell = ws.cell(row=row, column=6, value=m.confidence)
        conf_cell.number_format = "0.00"
        conf_cell.fill = conf_fill
        conf_cell.border = _border()
        conf_cell.alignment = _align("center")

    col_widths = [14, 30, 14, 35, 10, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


# ─── Mismatches 시트 ─────────────────────────────────────────────

def _write_mismatches(wb: Workbook, results: list[ReconcileResult]):
    ws = wb.create_sheet("Mismatches")
    ws.sheet_view.showGridLines = False

    headers = ["주석번호", "국문 주석제목", "국문 레이블", "영문 레이블",
               "속성(국문)", "국문 금액", "영문 금액", "차이", "상태", "신뢰도", "LLM 메모"]
    for col, h in enumerate(headers, 1):
        _write_cell(ws, 1, col, h,
                    fill=_fill(C_HEADER),
                    font=_font(bold=True, color="FFFFFF"),
                    align=_align("center"))

    row = 2
    for r in results:
        for item in r.items:
            if item.is_header_only:
                continue
            for am in item.amount_matches:
                if am.is_match is True:
                    continue  # 일치 항목 제외

                status = "미발견" if not am.found else "불일치"
                status_fill = _fill(C_NOT_FOUND if not am.found else C_MISMATCH)

                attr_str = ", ".join(f"{k}={v}" for k, v in am.attributes_kr.items() if not k.startswith("_"))

                _write_cell(ws, row, 1, r.note_number_kr)
                _write_cell(ws, row, 2, r.note_title_kr)
                _write_cell(ws, row, 3, item.label_kr)
                _write_cell(ws, row, 4, item.label_en or "—")
                _write_cell(ws, row, 5, attr_str)

                for col, val in enumerate([am.value_kr, am.value_en], 6):
                    c = _write_cell(ws, row, col, val, number_format="#,##0")
                    c.alignment = _align("right")

                # 차이 열: 엑셀 수식으로 표시
                if am.value_en is not None and am.value_kr is not None:
                    var_val = f"=G{row}-F{row}"
                else:
                    var_val = None
                c = _write_cell(ws, row, 8, var_val, number_format="#,##0")
                c.alignment = _align("right")

                _write_cell(ws, row, 9, status, fill=status_fill, align=_align("center"))

                conf_cell = ws.cell(row=row, column=10, value=am.confidence)
                conf_cell.number_format = "0.00"
                conf_cell.border = _border()
                conf_cell.alignment = _align("center")

                _write_cell(ws, row, 11, am.llm_note or "", align=_align(wrap=True))
                row += 1

    col_widths = [8, 25, 25, 25, 30, 14, 14, 14, 8, 7, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


# ─── Note_XX 시트 ────────────────────────────────────────────────

def _write_note_sheet(wb: Workbook, result: ReconcileResult, sheet_name: str | None = None):
    if sheet_name is None:
        try:
            sheet_name = f"Note_{int(result.note_number_kr):02d}"
        except (ValueError, TypeError):
            sheet_name = f"Note_{result.note_number_kr}"
    # Excel 시트명 31자 제한 및 금지문자 처리
    sheet_name = sheet_name[:31].replace("/", "_").replace("\\", "_").replace("?", "X")
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # ── 시트 제목 ───────────────────────────────────────────────
    ws.merge_cells("A1:B1")
    ws["A1"].value = f"주석 {result.note_number_kr}. {result.note_title_kr}"
    ws["A1"].font = _font(bold=True, color="FFFFFF", size=11)
    ws["A1"].fill = _fill(C_HEADER)
    ws["A1"].alignment = _align("left")

    ws.merge_cells("C1:D1")
    ws["C1"].value = f"Note {result.note_number_en or '?'}. {result.note_title_en or '—'}"
    ws["C1"].font = _font(bold=True, color="FFFFFF", size=11)
    ws["C1"].fill = _fill("2E7D32")
    ws["C1"].alignment = _align("left")

    # ── 다차원 속성 블록 구성 ────────────────────────────────────
    # 해당 주석 내 모든 AmountMatch의 속성 조합 수집
    # → 가장 많은 amount를 가진 item 기준으로 블록 수 결정
    max_amounts = max(
        (len(item.amount_matches) for item in result.items if not item.is_header_only),
        default=0,
    )

    if max_amounts == 0:
        # 금액 없는 주석
        ws["A2"].value = "금액 항목 없음 (텍스트 전용 주석)"
        ws["A2"].font = _font(color="888888")
        return

    # 모든 amount_matches에서 고유 속성 키 수집 (순서 보존)
    attr_keys_ordered = _collect_attr_keys(result)

    # 고정 열: 항목번호(A), 국문레이블(B), 영문레이블(C), 항목신뢰도(D) = 4열
    FIXED_COLS = 4
    # amount 블록 1개당 열 수: 속성열들 + 국문금액 + 영문금액 + 차이 + 일치 + 신뢰도 + 메모
    BLOCK_FIXED = 6  # 국문금액 + 영문금액 + 차이 + 일치 + 신뢰도 + 메모
    block_attr_count = len(attr_keys_ordered)
    block_width = block_attr_count + BLOCK_FIXED

    total_cols = FIXED_COLS + max_amounts * block_width

    # ── 헤더 행 (2행) ────────────────────────────────────────────
    HDR_ROW = 2
    # 고정 열 헤더
    fixed_headers = ["항목번호", "국문 레이블", "영문 레이블", "신뢰도(%)"]
    fixed_fills   = [C_HEADER,  C_KR_COL,    C_EN_COL,    C_ATTR]
    for col, (h, fc) in enumerate(zip(fixed_headers, fixed_fills), 1):
        _write_cell(ws, HDR_ROW, col, h,
                    fill=_fill(fc),
                    font=_font(bold=True, color="FFFFFF" if fc == C_HEADER else "000000"),
                    align=_align("center"))

    # 블록 헤더
    for blk_idx in range(max_amounts):
        start_col = FIXED_COLS + blk_idx * block_width + 1
        # 속성 열
        for ak_i, ak in enumerate(attr_keys_ordered):
            _write_cell(ws, HDR_ROW, start_col + ak_i,
                        ak, fill=_fill(C_ATTR), align=_align("center"))
        # 고정 블록 열
        block_fixed_headers = ["국문금액", "영문금액", "차이", "일치", "신뢰도", "LLM메모"]
        block_fills          = [C_KR_COL, C_EN_COL, "FCE4D6", "FFFFFF", C_ATTR, "F2F2F2"]
        for bfi, (bh, bf) in enumerate(zip(block_fixed_headers, block_fills)):
            _write_cell(ws, HDR_ROW, start_col + block_attr_count + bfi,
                        bh, fill=_fill(bf), align=_align("center"))

    # ── 데이터 행 ────────────────────────────────────────────────
    DATA_ROW = 3
    current_row = DATA_ROW

    for item in result.items:
        if item.is_header_only:
            # 제목행 — 컬럼 A~B 병합하여 표시
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row,   end_column=total_cols
            )
            c = ws.cell(row=current_row, column=1, value=item.label_kr)
            c.font = _font(bold=True, color="FFFFFF")
            c.fill = _fill("607D8B")
            c.alignment = _align("left")
            current_row += 1
            continue

        # 고정 열: 항목번호, 레이블, 영문레이블, 신뢰도
        avg_conf = (
            sum(am.confidence for am in item.amount_matches) / len(item.amount_matches)
            if item.amount_matches else 0.0
        )
        row_bg = _row_bg_fill(item.amount_matches)

        _write_cell(ws, current_row, 1, item.item_id,       fill=row_bg, align=_align("center"))
        _write_cell(ws, current_row, 2, item.label_kr,      fill=_fill(C_KR_COL))
        _write_cell(ws, current_row, 3, item.label_en or "—", fill=_fill(C_EN_COL))

        conf_cell = ws.cell(row=current_row, column=4, value=avg_conf)
        conf_cell.number_format = "0.0%"
        conf_cell.fill = _fill(C_ATTR)
        conf_cell.border = _border()
        conf_cell.alignment = _align("center")

        # 블록 열: 각 amount_match
        for blk_idx, am in enumerate(item.amount_matches):
            start_col = FIXED_COLS + blk_idx * block_width + 1

            # 속성 값들
            for ak_i, ak in enumerate(attr_keys_ordered):
                attr_val = am.attributes_kr.get(ak, "")
                _write_cell(ws, current_row, start_col + ak_i,
                            attr_val, fill=_fill(C_ATTR), align=_align("center"))

            # 금액 열들
            col_offset = start_col + block_attr_count
            cell_fill  = _amount_fill(am)

            kr_cell = ws.cell(row=current_row, column=col_offset, value=am.value_kr)
            kr_cell.number_format = "#,##0"
            kr_cell.fill = _fill(C_KR_COL)
            kr_cell.border = _border()
            kr_cell.alignment = _align("right")

            en_cell = ws.cell(row=current_row, column=col_offset + 1, value=am.value_en)
            en_cell.number_format = "#,##0"
            en_cell.fill = cell_fill
            en_cell.border = _border()
            en_cell.alignment = _align("right")

            # 차이 열: 영문 금액이 있는 경우 엑셀 수식으로 표시
            kr_col_ltr = get_column_letter(col_offset)
            en_col_ltr = get_column_letter(col_offset + 1)
            if am.value_en is not None and am.value_kr is not None:
                var_value = f"={en_col_ltr}{current_row}-{kr_col_ltr}{current_row}"
            else:
                var_value = None
            var_cell = ws.cell(row=current_row, column=col_offset + 2, value=var_value)
            var_cell.number_format = "#,##0"
            var_cell.fill = cell_fill
            var_cell.border = _border()
            var_cell.alignment = _align("right")

            match_icon = "✓" if am.is_match is True else ("✗" if am.is_match is False else "—")
            _write_cell(ws, current_row, col_offset + 3,
                        match_icon, fill=cell_fill, align=_align("center"))

            conf_am = ws.cell(row=current_row, column=col_offset + 4, value=am.confidence)
            conf_am.number_format = "0.00"
            conf_am.fill = cell_fill
            conf_am.border = _border()
            conf_am.alignment = _align("center")

            _write_cell(ws, current_row, col_offset + 5,
                        am.llm_note or "", align=_align(wrap=True))

        current_row += 1

    # ── 열 너비 조정 ─────────────────────────────────────────────
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 8

    for blk_idx in range(max_amounts):
        start_col = FIXED_COLS + blk_idx * block_width + 1
        for ak_i in range(block_attr_count):
            ws.column_dimensions[get_column_letter(start_col + ak_i)].width = 10
        ws.column_dimensions[get_column_letter(start_col + block_attr_count)].width = 14
        ws.column_dimensions[get_column_letter(start_col + block_attr_count + 1)].width = 14
        ws.column_dimensions[get_column_letter(start_col + block_attr_count + 2)].width = 12
        ws.column_dimensions[get_column_letter(start_col + block_attr_count + 3)].width = 5
        ws.column_dimensions[get_column_letter(start_col + block_attr_count + 4)].width = 6
        ws.column_dimensions[get_column_letter(start_col + block_attr_count + 5)].width = 25

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[HDR_ROW].height = 20
    ws.freeze_panes = f"A{DATA_ROW}"


# ─── 헬퍼 ───────────────────────────────────────────────────────

def _collect_attr_keys(result: ReconcileResult) -> list[str]:
    """
    결과 내 모든 amount_matches에서 사용된 속성 키를 순서 보존으로 수집.
    _is_pct 같은 내부 키는 제외.
    """
    seen: dict[str, None] = {}  # 순서 보존 set
    for item in result.items:
        for am in item.amount_matches:
            for k in am.attributes_kr:
                if not k.startswith("_"):
                    seen[k] = None
    return list(seen.keys())


def _row_bg_fill(amount_matches: list[AmountMatch]) -> PatternFill:
    """행 전체 배경: 하나라도 불일치면 연빨강, 모두 미발견이면 회색, 모두 일치면 흰색."""
    if not amount_matches:
        return _fill("FFFFFF")
    if any(am.is_match is False for am in amount_matches):
        return _fill("FFF0F0")
    if all(not am.found for am in amount_matches):
        return _fill("F5F5F5")
    return _fill("FFFFFF")


def _amount_fill(am: AmountMatch) -> PatternFill:
    """개별 금액 셀 배경색."""
    if not am.found:
        return _fill(C_NOT_FOUND)
    if am.is_match is True:
        return _fill(C_MATCH) if am.confidence >= 0.8 else _fill(C_LOW_CONF)
    if am.is_match is False:
        return _fill(C_MISMATCH)
    return _fill(C_LOW_CONF)
